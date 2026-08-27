#!/usr/bin/env python3
"""
OpenAlgo periodic security audit.

    uv run --with openpyxl python .claude/skills/security-audit/audit.py

Writes tmp/security_audit_<UTC-date>.xlsx. Everything it writes stays inside
tmp/, which is gitignored. The report NEVER contains secret values -- only
pass/fail, a redacted fingerprint, and a remediation pointer.

Read-only: it does not modify config, rotate keys, or touch the database.

Exit: 0 = no open CRITICAL/HIGH findings, 1 = at least one.
"""

from __future__ import annotations

import hashlib
import json
import os
import re
import stat
import subprocess
import sys
import tempfile
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path

REPO = Path(__file__).resolve().parents[3]
# Reports are organised under tmp/security-audits/<YYYY-MM>/ so a year of
# monthly runs stays browsable. tmp/ is gitignored and holds reports only.
NOW = datetime.now(timezone.utc)
TODAY = NOW.strftime("%Y-%m-%d")
OUTDIR = REPO / "tmp" / "security-audits" / NOW.strftime("%Y-%m")
TMP = OUTDIR
REPORT = OUTDIR / f"security_audit_{NOW.strftime('%Y-%m-%d_%H%M')}.xlsx"

# Publicly-known values that must never be live. From start.sh's preflight
# block -- every Docker install before v2.0.0.6 shipped these.
LEAKED = {
    "3daa0403ce2501ee7432b75bf100048e3cf510d63d2754f952729a991d8e2417",
    "a25d94718479b170c16278e321ea6c989358bf499a658fd20c90033cef8ce772",
    "OPENALGO_PLACEHOLDER_APP_KEY_REGENERATE_BEFORE_USE",
    "OPENALGO_PLACEHOLDER_API_KEY_PEPPER_REGENERATE_BEFORE_USE",
}

# Routes that are public BY DESIGN. Anything unprotected and not matching one
# of these is reported as drift -- that is the point of the check.
PUBLIC_ROUTE_ALLOWLIST = [
    r"^/$", r"^/login", r"^/setup", r"^/reset-password", r"^/download", r"^/faq",
    r"^/error", r"^/rate-limited", r"^/broker", r"^/favicon", r"^/logo", r"^/assets/",
    r"^/images/", r"^/sounds/", r"^/docs/", r"^/apple-touch-icon",
    r"^/csrf-token", r"^/broker-config", r"^/check-setup", r"^/session-status",
    r"^/app-info", r"^/logout", r"^/callback", r"/callback$", r"^/\.well-known",
    r"^/jwks\.json", r"^/healthz", r"^/postback/", r"^/webhook/", r"/webhook/",
    r"^/initiate-oauth", r"/auth$", r"/totp$",
]

findings: list[dict] = []


def add(check, severity, status, detail, remediation="", evidence=""):
    findings.append({
        "Check": check, "Severity": severity, "Status": status,
        "Detail": detail, "Remediation": remediation, "Evidence": evidence,
    })


def run(cmd, timeout=1200):
    try:
        p = subprocess.run(cmd, cwd=REPO, capture_output=True, text=True,
                           timeout=timeout, check=False)
        return p.returncode, p.stdout, p.stderr
    except subprocess.TimeoutExpired:
        return -1, "", "timeout"
    except FileNotFoundError:
        return -2, "", "tool not found"


def fingerprint(value: str) -> str:
    """Redacted, stable identifier for a secret -- safe to put in a report."""
    if not value:
        return "(empty)"
    return f"sha256:{hashlib.sha256(value.encode()).hexdigest()[:12]} len={len(value)}"


def truthy(v):
    return str(v).strip().lower() in ("true", "1", "t", "yes")


def parse_env() -> dict:
    env_path = REPO / ".env"
    if not env_path.exists():
        return {}
    out = {}
    for line in env_path.read_text(errors="replace").splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        k, _, v = line.partition("=")
        out[k.strip()] = v.strip().strip("'\"")
    return out


def pyfiles(*dirs):
    for d in dirs:
        root = REPO / d
        if root.exists():
            yield from root.rglob("*.py")


# ==========================================================================
# 1. Secrets and key hygiene
# ==========================================================================
def check_keys(env):
    if not env:
        add("Secrets: .env present", "HIGH", "WARN", ".env not found; key checks skipped",
            "Copy .sample.env to .env and generate fresh keys")
        return
    for key in ("APP_KEY", "API_KEY_PEPPER"):
        val = env.get(key, "")
        if not val:
            add(f"Secrets: {key} set", "CRITICAL", "FAIL", f"{key} missing from .env",
                'uv run python -c "import secrets; print(secrets.token_hex(32))"')
        elif val in LEAKED:
            add(f"Secrets: {key} not publicly known", "CRITICAL", "FAIL",
                f"{key} is a publicly-known leaked/placeholder value",
                "Rotate now. APP_KEY: safe (invalidates sessions). PEPPER: use "
                "upgrade/rotate_pepper.py -- never hand-edit on a populated DB.",
                fingerprint(val))
        elif len(val) < 64:
            add(f"Secrets: {key} entropy", "HIGH", "FAIL",
                f"{key} is {len(val)} chars; expected 64 hex (32 bytes)",
                "Regenerate with secrets.token_hex(32)", fingerprint(val))
        else:
            add(f"Secrets: {key} strength", "CRITICAL", "PASS",
                f"{key} unique and full length", "", fingerprint(val))


def check_secrets_not_committed():
    """Only real secret material counts -- placeholder readmes/.gitignore do not."""
    rc, out, _ = run(["git", "ls-files"])
    tracked = out.splitlines()
    bad = [f for f in tracked if re.match(r"^(\.env$|\.env\.|keys/.+|db/.+\.(db|duckdb)$)", f)
           and not f.endswith((".gitignore", ".sample", "readme.txt", "README.md"))]
    if bad:
        add("Secrets: no credentials tracked in git", "CRITICAL", "FAIL",
            f"Sensitive files tracked: {', '.join(bad[:6])}",
            "git rm --cached <path>; rotate anything exposed; purge history with git-filter-repo")
    else:
        add("Secrets: no credentials tracked in git", "CRITICAL", "PASS",
            ".env, keys/* and database files are untracked")


def check_detect_secrets():
    rc, out, err = run([
        "uv", "run", "--group", "dev", "detect-secrets", "scan", "--exclude-files",
        r"(\.venv|frontend/(node_modules|dist)|tmp/|db/|log/|uv\.lock|package-lock\.json)",
    ])
    if rc < 0:
        add("Secrets: detect-secrets scan", "HIGH", "ERROR", f"could not run ({err})",
            "uv sync --group dev")
        return
    try:
        results = json.loads(out).get("results", {})
    except json.JSONDecodeError:
        add("Secrets: detect-secrets scan", "HIGH", "ERROR", "unparseable output")
        return
    total = sum(len(v) for v in results.values())
    if total:
        add("Secrets: no plaintext secrets in tree", "HIGH", "REVIEW",
            f"{total} candidates across {len(results)} files",
            "Triage each. Real leak -> rotate AND purge from git history; "
            "false positive -> add to .secrets.baseline",
            "; ".join(f"{f} ({len(v)})" for f, v in list(results.items())[:6]))
    else:
        add("Secrets: no plaintext secrets in tree", "HIGH", "PASS", "no candidates")


# ==========================================================================
# 2. Runtime posture
# ==========================================================================
def check_runtime_config(env):
    debug = truthy(env.get("FLASK_DEBUG", "False"))
    add("Runtime: FLASK_DEBUG disabled", "CRITICAL", "FAIL" if debug else "PASS",
        "FLASK_DEBUG ON -- Werkzeug console allows RCE" if debug else "FLASK_DEBUG off",
        "Set FLASK_DEBUG = 'False' before exposing the instance")

    mcp = truthy(env.get("MCP_HTTP_ENABLED", "False"))
    add("Runtime: remote MCP not in debug", "CRITICAL",
        "FAIL" if (mcp and debug) else "PASS",
        f"MCP_HTTP_ENABLED={mcp}, FLASK_DEBUG={debug}",
        "app.py refuses this combination at startup; disable one" if (mcp and debug) else "")

    add("Runtime: CSRF enabled", "HIGH",
        "PASS" if truthy(env.get("CSRF_ENABLED", "TRUE")) else "FAIL",
        f"CSRF_ENABLED={env.get('CSRF_ENABLED', 'TRUE')}",
        "Never disable outside local development")

    csp, csp_ro = env.get("CSP_ENABLED", "TRUE"), env.get("CSP_REPORT_ONLY", "FALSE")
    if not truthy(csp):
        add("Runtime: CSP enforced", "HIGH", "FAIL", "CSP disabled", "Set CSP_ENABLED = 'TRUE'")
    elif truthy(csp_ro):
        add("Runtime: CSP enforced", "MEDIUM", "WARN",
            "CSP report-only -- violations logged, not blocked",
            "Set CSP_REPORT_ONLY = 'FALSE' once tuned")
    else:
        add("Runtime: CSP enforced", "HIGH", "PASS", "CSP enabled and enforcing")

    origins = env.get("CORS_ALLOWED_ORIGINS", "")
    creds = truthy(env.get("CORS_ALLOW_CREDENTIALS", "FALSE"))
    if "*" in origins and creds:
        add("Runtime: CORS not wildcard+credentials", "CRITICAL", "FAIL",
            "CORS '*' with credentials -- any site can call the API as the user",
            "Pin CORS_ALLOWED_ORIGINS to your exact host")
    elif "*" in origins:
        add("Runtime: CORS not wildcard+credentials", "MEDIUM", "WARN",
            "CORS origin '*' (credentials off)", "Pin to your exact host")
    else:
        add("Runtime: CORS not wildcard+credentials", "HIGH", "PASS", "CORS origins pinned")

    host = env.get("HOST_SERVER", "")
    local = ("127.0.0.1" in host) or ("localhost" in host)
    if host.startswith("http://") and not local:
        add("Runtime: HTTPS in use", "HIGH", "FAIL",
            "HOST_SERVER is plain http on a non-loopback host -- session cookies "
            "lose the Secure flag and traffic is unencrypted",
            "Terminate TLS (install/ nginx setup) and set HOST_SERVER to https://")
    else:
        add("Runtime: HTTPS in use", "HIGH", "PASS", "HOST_SERVER scheme acceptable")

    if truthy(env.get("NGROK_ALLOW", "FALSE")):
        add("Runtime: ngrok tunnel off", "MEDIUM", "WARN",
            "NGROK_ALLOW enabled -- exposes the instance via a public tunnel",
            "Disable unless actively debugging")
    else:
        add("Runtime: ngrok tunnel off", "MEDIUM", "PASS", "NGROK_ALLOW off")

    missing = [k for k in ("LOGIN_RATE_LIMIT_MIN", "LOGIN_RATE_LIMIT_HOUR",
                           "RESET_RATE_LIMIT", "API_RATE_LIMIT", "ORDER_RATE_LIMIT")
               if not env.get(k)]
    add("Runtime: rate limits configured", "MEDIUM", "WARN" if missing else "PASS",
        f"unset (using defaults): {', '.join(missing)}" if missing
        else "login, reset, api and order limits all set",
        "Set explicitly in .env so limits survive a config reset" if missing else "")


# ==========================================================================
# 3. Route protection
# ==========================================================================
def check_route_protection():
    rows, total = [], 0
    for py in sorted((REPO / "blueprints").glob("*.py")):
        src = py.read_text(errors="replace").splitlines()
        for i, line in enumerate(src):
            if not re.match(r"\s*@(\w+_bp)\.route\(", line):
                continue
            total += 1
            path_m = re.search(r'route\(\s*["\']([^"\']*)["\']', line)
            path = path_m.group(1) if path_m else "?"
            decs, j = [], i
            while j < len(src) and not re.match(r"\s*(async )?def ", src[j]):
                decs.append(src[j]); j += 1
            blob = "\n".join(decs)
            # Authentication only. A rate limit throttles an attacker; it does
            # not authenticate them, so limiter.limit alone is NOT protection.
            if re.search(r"check_session_validity", blob):
                continue
            if any(re.search(p, path) for p in PUBLIC_ROUTE_ALLOWLIST):
                continue
            methods = re.search(r"methods=\[([^\]]*)\]", line)
            rows.append({
                "file": str(py.relative_to(REPO)), "line": i + 1, "path": path,
                "methods": (methods.group(1).replace('"', "").replace("'", "")
                            if methods else "GET"),
                "handler": src[j].strip()[:60] if j < len(src) else "?",
            })
    state_changing = [r for r in rows if any(m in r["methods"] for m in ("POST", "PUT", "DELETE", "PATCH"))]
    if state_changing:
        add("Routes: no unprotected state-changing endpoints", "CRITICAL", "REVIEW",
            f"{len(state_changing)} of {total} routes accept writes without "
            "@check_session_validity or a rate limit, and are not on the public allowlist",
            "Add @check_session_validity, or add to PUBLIC_ROUTE_ALLOWLIST with a reason",
            "; ".join(f"{r['file']}:{r['line']} {r['path']}" for r in state_changing[:6]))
    else:
        add("Routes: no unprotected state-changing endpoints", "CRITICAL", "PASS",
            f"all write routes of {total} are protected or explicitly public")

    read_only = [r for r in rows if r not in state_changing]
    if read_only:
        add("Routes: unprotected read endpoints reviewed", "MEDIUM", "REVIEW",
            f"{len(read_only)} read-only routes are public and not on the allowlist",
            "Confirm none leak account data; add legitimate ones to the allowlist",
            "; ".join(f"{r['file']}:{r['line']} {r['path']}" for r in read_only[:8]))
    else:
        add("Routes: unprotected read endpoints reviewed", "MEDIUM", "PASS",
            "no unexpected public read routes")
    return rows


def check_dev_routes():
    """Test/debug surfaces should not be registered in a production app."""
    hits = []
    for py in (REPO / "blueprints").glob("*.py"):
        src = py.read_text(errors="replace")
        for m in re.finditer(r'route\(\s*["\']([^"\']*(?:test|debug|dev)[^"\']*)["\']', src):
            line = src[:m.start()].count("\n") + 1
            hits.append(f"{py.relative_to(REPO)}:{line} {m.group(1)}")
    if hits:
        add("Routes: no test/debug endpoints registered", "MEDIUM", "REVIEW",
            f"{len(hits)} route(s) look like test or debug surfaces",
            "Gate behind FLASK_DEBUG or remove the blueprint registration in app.py",
            "; ".join(hits[:6]))
    else:
        add("Routes: no test/debug endpoints registered", "MEDIUM", "PASS", "none found")


# ==========================================================================
# 4. Database security
# ==========================================================================
def check_database(env):
    dbdir = REPO / "db"
    if dbdir.exists():
        world = [p.name for p in dbdir.glob("*")
                 if p.is_file() and stat.S_IMODE(p.stat().st_mode) & 0o007]
        add("Database: files not world-accessible", "HIGH",
            "FAIL" if world else "PASS",
            f"world-accessible: {', '.join(world[:6])}" if world
            else f"{len(list(dbdir.glob('*.db')))} database file(s), owner/group only",
            "chmod 640 db/*.db && chmod 750 db" if world else "")
    else:
        add("Database: files not world-accessible", "HIGH", "SKIP", "db/ not present")

    src = (REPO / "database" / "auth_db.py")
    if src.exists():
        text = src.read_text(errors="replace")
        add("Database: broker tokens encrypted at rest", "CRITICAL",
            "PASS" if "Fernet" in text else "FAIL",
            "Fernet encryption present in auth_db" if "Fernet" in text
            else "Fernet not found -- tokens may be stored in plaintext",
            "" if "Fernet" in text else "Investigate immediately")
        peppered = re.search(r"PEPPER|pepper", text) is not None
        add("Database: API keys peppered before storage", "CRITICAL",
            "PASS" if peppered else "FAIL",
            "pepper referenced in auth_db" if peppered else "no pepper reference found")

    # NullPool is a correctness AND availability control (see CLAUDE.md).
    ef = REPO / "database" / "engine_factory.py"
    if ef.exists():
        # Match how the pool is actually configured, not any mention of it. A
        # plain `"StaticPool" not in text` test fails on this file's own
        # docstring, which exists to say StaticPool must never be used: the
        # check reported the documentation as the defect that documentation
        # warns against. Look for the assignment and the import instead, with
        # line comments stripped so a commented-out example cannot trip it.
        code = "\n".join(
            ln for ln in ef.read_text(errors="replace").splitlines()
            if not ln.lstrip().startswith("#")
        )
        uses_null = re.search(r"poolclass\s*=\s*NullPool", code) is not None
        uses_static = bool(
            re.search(r"poolclass\s*=\s*StaticPool", code)
            or re.search(r"^\s*from\s+sqlalchemy\.pool\s+import\s+[^\n]*StaticPool", code, re.M)
        )
        ok = uses_null and not uses_static
        add("Database: NullPool (not StaticPool)", "MEDIUM", "PASS" if ok else "FAIL",
            "engine_factory configures poolclass=NullPool" if ok
            else "StaticPool configured -- corrupts cursor state under concurrency",
            "" if ok else "Revert to NullPool; see CLAUDE.md")

    # Direct create_engine calls bypass the hardened factory.
    strays = []
    for py in pyfiles("database", "services", "blueprints", "broker"):
        try:
            for i, line in enumerate(py.read_text(errors="replace").splitlines(), 1):
                if re.search(r"\bcreate_engine\(", line) and "engine_factory" not in str(py):
                    strays.append(f"{py.relative_to(REPO)}:{i}")
        except OSError:
            continue
    add("Database: engines created via factory", "MEDIUM",
        "REVIEW" if strays else "PASS",
        f"{len(strays)} direct create_engine call(s) bypassing engine_factory" if strays
        else "no direct create_engine calls",
        "Route through database.engine_factory.create_db_engine() for NullPool + FD hygiene",
        "; ".join(strays[:6]))

    backups = list((REPO / "db").glob("*.bak")) + list(REPO.glob("*backup*"))
    add("Database: backup present", "MEDIUM", "PASS" if backups else "WARN",
        f"{len(backups)} backup artifact(s) found" if backups
        else "no backup artifacts in the repo tree",
        "Confirm backups exist off-box and are stored encrypted -- db/openalgo.db "
        "holds password hashes and encrypted broker tokens")


def check_csrf_exemptions():
    """Every csrf.exempt is a deliberate hole. Each needs its own auth story."""
    app_py = REPO / "app.py"
    if not app_py.exists():
        add("Routes: CSRF exemptions justified", "HIGH", "SKIP", "app.py not found")
        return
    lines = app_py.read_text(errors="replace").splitlines()
    exempts = [f"{i}: {ln.strip()[:90]}" for i, ln in enumerate(lines, 1)
               if re.search(r"csrf\.exempt\(", ln)]
    if exempts:
        add("Routes: CSRF exemptions justified", "HIGH", "REVIEW",
            f"{len(exempts)} CSRF exemption(s) in app.py",
            "Each exemption MUST carry its own authentication: API-key auth for "
            "/api/v1, a bearer token in the URL for webhooks, or broker-signed "
            "data for postbacks. An exemption on a session-authenticated route "
            "is a CSRF vulnerability.",
            "; ".join(exempts[:8]))
    else:
        add("Routes: CSRF exemptions justified", "HIGH", "PASS", "no exemptions")


def check_security_headers():
    """Headers beyond CSP. HSTS lives in nginx, the rest in csp.py."""
    csp_py = REPO / "csp.py"
    required = {
        "X-Frame-Options": "HIGH",
        "X-Content-Type-Options": "MEDIUM",
        "Referrer-Policy": "MEDIUM",
        "Permissions-Policy": "LOW",
    }
    text = csp_py.read_text(errors="replace") if csp_py.exists() else ""
    missing = [h for h in required if h not in text]
    if missing:
        add("Headers: browser hardening headers set", "HIGH", "FAIL",
            f"missing from csp.py: {', '.join(missing)}",
            "Add to the header block in csp.py")
    else:
        add("Headers: browser hardening headers set", "HIGH", "PASS",
            "X-Frame-Options, X-Content-Type-Options, Referrer-Policy, Permissions-Policy all set")

    # HSTS belongs at the TLS terminator, not the app.
    hsts = any("Strict-Transport-Security" in p.read_text(errors="replace")
               for p in (REPO / "install").glob("*.sh") if p.is_file())
    add("Headers: HSTS configured at TLS terminator", "MEDIUM",
        "PASS" if hsts else "WARN",
        "Strict-Transport-Security present in the nginx templates" if hsts
        else "no HSTS found in install/ nginx templates",
        "" if hsts else "Add 'add_header Strict-Transport-Security \"max-age=63072000\" always;'")


def check_auth_hardening():
    """Password storage, session lifetime, and brute-force resistance."""
    hashers = []
    for py in pyfiles("database", "utils"):
        try:
            t = py.read_text(errors="replace")
        except OSError:
            continue
        if re.search(r"from argon2 import|PasswordHasher", t):
            hashers.append(str(py.relative_to(REPO)))
        if re.search(r"hashlib\.(md5|sha1|sha256)\(.*password", t, re.I):
            add("Auth: no bare-hash password storage", "CRITICAL", "FAIL",
                f"{py.relative_to(REPO)} appears to hash a password with a bare digest",
                "Passwords must use Argon2 (PasswordHasher), never a raw hash")
    add("Auth: Argon2 password hashing", "CRITICAL", "PASS" if hashers else "FAIL",
        f"Argon2 PasswordHasher used in {len(hashers)} module(s)" if hashers
        else "no Argon2 usage found",
        "" if hashers else "Investigate -- password hashing may have been downgraded",
        ", ".join(hashers[:5]))

    # API keys must be hashed+peppered at rest, never stored raw.
    ad = REPO / "database" / "auth_db.py"
    if ad.exists():
        t = ad.read_text(errors="replace")
        ok = re.search(r"PasswordHasher|argon2", t) and re.search(r"pepper", t, re.I)
        add("Auth: API keys hashed with pepper at rest", "CRITICAL",
            "PASS" if ok else "FAIL",
            "auth_db hashes API keys with a pepper" if ok
            else "API key hashing/peppering not detected in auth_db",
            "" if ok else "API keys must never be recoverable from the database")

    sess = REPO / "utils" / "session.py"
    if sess.exists() and re.search(r"session\.clear\(\)|regenerate", sess.read_text(errors="replace")):
        add("Auth: session invalidated on logout", "HIGH", "PASS", "session teardown present")
    else:
        add("Auth: session invalidated on logout", "HIGH", "REVIEW",
            "no explicit session.clear() found in utils/session.py",
            "Confirm logout fully clears the server-side session")

    # Login brute-force resistance is rate limiting, not lockout, in this app.
    bl = REPO / "blueprints" / "auth.py"
    if bl.exists():
        t = bl.read_text(errors="replace")
        limited = t.count("@limiter.limit")
        add("Auth: login endpoints rate limited", "HIGH",
            "PASS" if limited >= 2 else "REVIEW",
            f"{limited} rate-limited handler(s) in blueprints/auth.py",
            "Login and password-reset must both be limited -- they are the "
            "brute-force surface for a single-user deployment")


def check_strategy_sandbox():
    """The Python strategy host executes user-supplied code by design.
    That is intended, so the controls are isolation and FD hygiene, not
    prevention."""
    ps = REPO / "blueprints" / "python_strategy.py"
    if not ps.exists():
        add("Strategy host: subprocess isolation", "HIGH", "SKIP", "python_strategy.py not found")
        return
    t = ps.read_text(errors="replace")

    add("Strategy host: runs code in a subprocess", "HIGH",
        "PASS" if "subprocess" in t and "Popen" in t else "FAIL",
        "strategies execute in a separate process" if "Popen" in t
        else "no subprocess isolation detected",
        "In-process exec of user strategies would give them the Flask app's "
        "memory, including decrypted broker tokens")

    add("Strategy host: no shell=True", "CRITICAL",
        "FAIL" if re.search(r"shell\s*=\s*True", t) else "PASS",
        "shell=True present -- strategy names could reach a shell"
        if re.search(r"shell\s*=\s*True", t) else "no shell=True",
        "Pass an argv list, never a shell string")

    add("Strategy host: uses sys.executable", "MEDIUM",
        "PASS" if "sys.executable" in t else "REVIEW",
        "interpreter resolved via sys.executable" if "sys.executable" in t
        else "interpreter path may come from PATH",
        "A PATH-resolved interpreter can be hijacked")

    # PIPE without a reader deadlocks; CLAUDE.md requires a log file.
    if re.search(r'"stdout":\s*subprocess\.PIPE', t):
        add("Strategy host: subprocess output not on an unread PIPE", "MEDIUM", "REVIEW",
            "stdout is a PIPE -- confirm a reader thread drains it",
            "An undrained PIPE deadlocks the strategy once the buffer fills; "
            "the convention is to write to a log file and .wait()-reap")
    else:
        add("Strategy host: subprocess output not on an unread PIPE", "MEDIUM", "PASS",
            "stdout not an unread PIPE")

    # Path traversal into the strategy directory.
    traversal = re.search(r"os\.path\.join\([^)]*request\.|secure_filename", t)
    add("Strategy host: filename traversal guarded", "HIGH",
        "PASS" if "secure_filename" in t else "REVIEW",
        "secure_filename used" if "secure_filename" in t
        else "no secure_filename -- verify strategy filenames cannot traverse",
        "A '../' in a strategy name must not escape the strategies directory")


def check_websocket_auth():
    srv = REPO / "websocket_proxy" / "server.py"
    if not srv.exists():
        add("WebSocket: client authentication", "HIGH", "SKIP", "server.py not found")
        return
    t = srv.read_text(errors="replace")
    add("WebSocket: client authentication required", "CRITICAL",
        "PASS" if "verify_api_key" in t else "FAIL",
        "proxy verifies an API key before streaming" if "verify_api_key" in t
        else "no API-key verification found on the WS proxy",
        "" if "verify_api_key" in t else "Port 8765 would stream market data to anyone")
    grace = re.search(r"unauthenticated|grace", t, re.I)
    add("WebSocket: unauthenticated connections timed out", "MEDIUM",
        "PASS" if grace else "REVIEW",
        "an unauthenticated grace window is enforced" if grace
        else "no grace-window logic detected",
        "Without it, unauthenticated sockets can exhaust FDs on a public port")


def check_container_hardening():
    df = REPO / "Dockerfile"
    if not df.exists():
        add("Container: runs as non-root", "HIGH", "SKIP", "no Dockerfile")
        return
    t = df.read_text(errors="replace")
    add("Container: runs as non-root", "HIGH", "PASS" if re.search(r"^USER\s+(?!root)", t, re.M) else "FAIL",
        "USER directive drops from root" if re.search(r"^USER\s+(?!root)", t, re.M)
        else "no non-root USER -- container runs as root",
        "Add a non-root USER before CMD")
    secrets_in_image = re.search(r"^(ENV|ARG)\s+.*(SECRET|PASSWORD|API_KEY|TOKEN)", t, re.M | re.I)
    add("Container: no secrets baked into image", "CRITICAL",
        "FAIL" if secrets_in_image else "PASS",
        "a secret-looking ENV/ARG is set in the Dockerfile" if secrets_in_image
        else "no secret ENV/ARG in the Dockerfile",
        "Image layers are readable by anyone who pulls the image" if secrets_in_image else "")
    add("Container: .env not copied into image", "HIGH",
        "FAIL" if re.search(r"^COPY\s+.*\.env(\s|$)", t, re.M) else "PASS",
        ".env is COPYed into the image" if re.search(r"^COPY\s+.*\.env(\s|$)", t, re.M)
        else ".env is mounted, not baked in")


def check_git_history_secrets():
    """The working tree can be clean while history still leaks."""
    rc, out, _ = run(["git", "log", "--all", "--full-history", "--name-only",
                      "--pretty=format:", "--", ".env", "keys/*", "*.pem", "*.key"])
    leaked = sorted({ln.strip() for ln in out.splitlines() if ln.strip()
                     and not ln.strip().endswith((".gitignore", ".sample", "README.md", "readme.txt"))})
    if leaked:
        add("Secrets: none in git history", "CRITICAL", "REVIEW",
            f"{len(leaked)} secret-shaped path(s) appear in git history: {', '.join(leaked[:5])}",
            "If any real credential was ever committed, rotate it FIRST, then "
            "purge with git-filter-repo and force-push. Removing the file in a "
            "later commit does not remove it from history.")
    else:
        add("Secrets: none in git history", "CRITICAL", "PASS",
            "no .env/keys/pem/key paths in git history")


def check_logging_hygiene():
    logdir = REPO / "log"
    if logdir.exists():
        world = [p.name for p in logdir.glob("*")
                 if p.is_file() and stat.S_IMODE(p.stat().st_mode) & 0o007]
        add("Logging: log files not world-readable", "MEDIUM",
            "FAIL" if world else "PASS",
            f"world-readable: {', '.join(world[:5])}" if world else "log files owner/group only",
            "chmod 640 log/*" if world else "")
    else:
        add("Logging: log files not world-readable", "MEDIUM", "SKIP", "log/ not present")

    lg = REPO / "utils" / "logging.py"
    if lg.exists():
        t = lg.read_text(errors="replace")
        patterns = len(re.findall(r"(api[_-]?key|token|secret|password|pepper)", t, re.I))
        add("Logging: sensitive-data redaction coverage", "HIGH",
            "PASS" if patterns >= 4 else "REVIEW",
            f"SensitiveDataFilter references {patterns} secret-shaped term(s)",
            "The filter must cover api key, token, secret and password at minimum -- "
            "broker tokens flow through debug logs during auth")


def check_dependency_integrity():
    lock = REPO / "uv.lock"
    add("Dependencies: lockfile present", "HIGH", "PASS" if lock.exists() else "FAIL",
        "uv.lock present -- installs are reproducible" if lock.exists()
        else "no uv.lock; installs are not reproducible",
        "Run uv sync to generate it")
    pyproject = REPO / "pyproject.toml"
    if pyproject.exists():
        t = pyproject.read_text(errors="replace")
        alt_index = re.search(r"(index-url|extra-index-url)\s*=", t)
        add("Dependencies: no alternate package index", "HIGH",
            "REVIEW" if alt_index else "PASS",
            "an alternate index-url is configured" if alt_index
            else "packages come from PyPI only",
            "An alternate index is a supply-chain path; confirm it is intentional and trusted")
        unpinned = len(re.findall(r'"[a-zA-Z0-9_.-]+"\s*,', t))
        loose = len(re.findall(r'">=', t))
        add("Dependencies: version pinning", "MEDIUM", "REVIEW" if loose > 20 else "PASS",
            f"{loose} dependencies use a floating >= constraint",
            "Floating constraints let a compromised upstream release reach you on "
            "the next sync; uv.lock mitigates this as long as it is committed")


# ==========================================================================
# 4b. Cache security
# ==========================================================================
def check_cache():
    """Caches hold decrypted credentials in process memory -- TTL and
    invalidation are security controls, not performance tuning."""
    sensitive = []
    unbounded = []
    for py in pyfiles("database", "services", "utils", "broker"):
        try:
            lines = py.read_text(errors="replace").splitlines()
        except OSError:
            continue
        for i, line in enumerate(lines, 1):
            m = re.search(r"(\w*cache\w*)\s*=\s*TTLCache\((.*)\)", line, re.I)
            if m:
                name, args = m.group(1), m.group(2)
                is_sensitive = re.search(r"cred|token|auth|key|secret|password|session", name, re.I)
                if "ttl=" not in args:
                    unbounded.append(f"{py.relative_to(REPO)}:{i} {name} (no ttl)")
                if is_sensitive:
                    ttl = re.search(r"ttl\s*=\s*(\d+)", args)
                    sensitive.append(f"{py.relative_to(REPO)}:{i} {name} "
                                     f"ttl={ttl.group(1) + 's' if ttl else 'NONE'}")
            # A plain dict used as a credential cache never expires.
            if re.search(r"^_?\w*(cred|token|auth|secret)\w*cache\w*\s*=\s*\{\}", line, re.I):
                unbounded.append(f"{py.relative_to(REPO)}:{i} plain-dict credential cache")

    if unbounded:
        add("Cache: no unbounded credential caches", "HIGH", "REVIEW",
            f"{len(unbounded)} cache(s) without a TTL",
            "Every cache holding auth material needs an explicit ttl= so a revoked "
            "credential cannot live forever in process memory",
            "; ".join(unbounded[:6]))
    else:
        add("Cache: no unbounded credential caches", "HIGH", "PASS",
            "all caches have an explicit TTL")

    if sensitive:
        add("Cache: sensitive caches have bounded TTL", "MEDIUM", "REVIEW",
            f"{len(sensitive)} cache(s) hold auth-related data",
            "Confirm each TTL is acceptable and that logout/revoke invalidates it. "
            "Broker tokens roll at ~3 AM IST, so a TTL beyond that is a stale-auth risk.",
            "; ".join(sensitive[:8]))
    else:
        add("Cache: sensitive caches have bounded TTL", "MEDIUM", "PASS",
            "no auth-related caches detected")

    ci = REPO / "database" / "cache_invalidation.py"
    if ci.exists():
        t = ci.read_text(errors="replace")
        ok = "publish_auth_cache_invalidation" in t and "publish_all_cache_invalidation" in t
        add("Cache: auth-change invalidation available", "HIGH", "PASS" if ok else "FAIL",
            "cache_invalidation exposes auth and all-cache publishers" if ok
            else "invalidation publishers missing",
            "" if ok else "Without invalidation, a revoked token stays live in the proxy")
        # The ZMQ bus invariant is also a security control: a publisher that
        # binds can silently steal the port and stop invalidations arriving.
        binds = re.search(r"\.bind\(", t)
        add("Cache: invalidation publisher connects (never binds)", "HIGH",
            "FAIL" if binds else "PASS",
            "publisher calls bind() -- races the proxy SUB and silently drops "
            "invalidations" if binds else "publisher uses connect()",
            "See the ZeroMQ bus invariant in CLAUDE.md" if binds else "")
    else:
        add("Cache: auth-change invalidation available", "HIGH", "SKIP",
            "database/cache_invalidation.py not present")


# ==========================================================================
# 4c. Frontend security
# ==========================================================================
def check_frontend():
    fe = REPO / "frontend"
    if not (fe / "package.json").exists():
        add("Frontend: present", "LOW", "SKIP", "frontend/ not found")
        return []

    rows = []
    src = fe / "src"
    patterns = [
        ("XSS: dangerouslySetInnerHTML", r"dangerouslySetInnerHTML", "HIGH"),
        ("XSS: direct innerHTML assignment", r"\.innerHTML\s*=", "HIGH"),
        ("RCE: eval / new Function", r"\beval\s*\(|new\s+Function\s*\(", "HIGH"),
        ("Secrets: credential in browser storage",
         r"(localStorage|sessionStorage)\.setItem\(\s*[\"'][^\"']*"
         r"(token|key|secret|password|apikey|api_key|auth)", "HIGH"),
        ("Secrets: hardcoded key literal",
         r"(api[_-]?key|secret|password)\s*[:=]\s*[\"'][A-Za-z0-9]{16,}[\"']", "CRITICAL"),
        ("Transport: hardcoded http:// endpoint", r"[\"']http://(?!localhost|127\.0\.0\.1)", "MEDIUM"),
    ]
    for name, pat, sev in patterns:
        hits = []
        for f in list(src.rglob("*.ts")) + list(src.rglob("*.tsx")):
            try:
                for i, line in enumerate(f.read_text(errors="replace").splitlines(), 1):
                    if line.lstrip().startswith(("//", "*")):
                        continue
                    if re.search(pat, line, re.I):
                        rel = f"{f.relative_to(REPO)}:{i}"
                        hits.append(rel)
                        rows.append([sev, name, str(f.relative_to(REPO)), i, line.strip()[:150]])
            except OSError:
                continue
        if hits:
            add(f"Frontend: {name}", sev, "REVIEW", f"{len(hits)} occurrence(s)",
                "Confirm the value is developer-controlled and never user or API data. "
                "See the 'Frontend patterns' sheet.",
                "; ".join(hits[:5]) + (" ..." if len(hits) > 5 else ""))
        else:
            add(f"Frontend: {name}", sev, "PASS", "no occurrences")

    # Source maps in a shipped build leak original sources.
    dist = fe / "dist"
    if dist.exists():
        maps = list(dist.rglob("*.map"))
        add("Frontend: no source maps in dist", "MEDIUM", "REVIEW" if maps else "PASS",
            f"{len(maps)} .map file(s) in the shipped build" if maps
            else "no source maps shipped",
            "Disable sourcemap generation for production builds" if maps else "")
    else:
        add("Frontend: no source maps in dist", "MEDIUM", "SKIP", "frontend/dist not built")

    # npm audit -- JS dependency CVEs (separate ecosystem from pip-audit).
    rc, out, err = run(["npm", "audit", "--json", "--prefix", str(fe)], timeout=600)
    if rc == -2:
        add("Frontend: npm dependency CVEs", "HIGH", "SKIP", "npm not installed on this machine")
    else:
        try:
            meta = json.loads(out).get("metadata", {}).get("vulnerabilities", {})
            crit = meta.get("critical", 0) + meta.get("high", 0)
            tot = sum(v for k, v in meta.items() if k != "total")
            add("Frontend: npm dependency CVEs", "HIGH",
                "FAIL" if crit else ("REVIEW" if tot else "PASS"),
                f"{tot} advisories (critical/high: {crit})",
                "cd frontend && npm audit fix; review breaking changes before --force",
                json.dumps(meta))
        except (json.JSONDecodeError, AttributeError):
            add("Frontend: npm dependency CVEs", "HIGH", "ERROR", "npm audit output unparseable")
    return rows


# ==========================================================================
# 5. Code vulnerability patterns
# ==========================================================================
CODE_PATTERNS = [
    ("RCE: eval/exec", r"\b(eval|exec)\s*\(", "CRITICAL",
     ("blueprints", "restx_api", "services", "utils")),
    ("RCE: shell=True", r"shell\s*=\s*True", "CRITICAL",
     ("blueprints", "services", "utils", "broker")),
    ("Deserialization: pickle", r"pickle\.loads?\(", "HIGH",
     ("blueprints", "services", "database", "utils")),
    ("Deserialization: yaml.load", r"yaml\.load\((?![^)]*SafeLoader)", "HIGH",
     ("blueprints", "services", "utils")),
    ("SQLi: raw SQL f-string", r"(text|execute)\(\s*f[\"']", "HIGH",
     ("database", "services", "blueprints")),
    ("Path traversal: unvalidated join", r"os\.path\.join\([^)]*request\.", "HIGH",
     ("blueprints", "services")),
    ("Path traversal: send_file on input", r"send_file\([^)]*request\.", "HIGH",
     ("blueprints",)),
    ("SSRF: request-derived URL fetch", r"(requests|httpx|urlopen)\S*\([^)]*request\.(args|form|json)", "HIGH",
     ("blueprints", "services")),
    ("Open redirect: redirect on input", r"redirect\(\s*request\.(args|form)", "MEDIUM",
     ("blueprints",)),
    ("Crypto: weak hash", r"hashlib\.(md5|sha1)\((?![^)]*usedforsecurity)", "MEDIUM",
     ("blueprints", "services", "database", "utils", "broker")),
    ("Transport: TLS verification disabled", r"verify\s*=\s*False", "CRITICAL",
     ("broker", "services", "utils", "blueprints")),
    ("Availability: HTTP call without timeout",
     r"(httpx|requests)\.(get|post|put|delete)\((?![^)]*timeout)", "MEDIUM",
     ("broker", "services", "utils")),
    ("Logging: traceback bypasses central logging", r"traceback\.(print_exc|format_exc)", "LOW",
     ("blueprints", "services", "broker", "utils", "database")),
    ("Logging: print() instead of logger", r"^\s*print\(", "LOW",
     ("blueprints", "services", "restx_api")),
]


def check_code_patterns():
    detail_rows = []
    for name, pat, sev, dirs in CODE_PATTERNS:
        hits = []
        for py in pyfiles(*dirs):
            try:
                for i, line in enumerate(py.read_text(errors="replace").splitlines(), 1):
                    if line.lstrip().startswith("#"):
                        continue
                    if re.search(pat, line):
                        rel = f"{py.relative_to(REPO)}:{i}"
                        hits.append(rel)
                        detail_rows.append([sev, name, str(py.relative_to(REPO)), i,
                                            line.strip()[:150]])
            except OSError:
                continue
        if hits:
            add(f"Code: {name}", sev, "REVIEW", f"{len(hits)} occurrence(s)",
                "Confirm each: interpolated values must be developer-controlled constants, "
                "never request data. See the 'Code patterns' sheet.",
                "; ".join(hits[:5]) + (" ..." if len(hits) > 5 else ""))
        else:
            add(f"Code: {name}", sev, "PASS", "no occurrences")
    return detail_rows


# ==========================================================================
# 6. Application invariants
# ==========================================================================
def check_app_invariants():
    checks = [
        ("App: session cookie HttpOnly", "app.py", r"SESSION_COOKIE_HTTPONLY\s*=\s*True", "HIGH"),
        ("App: session cookie SameSite", "app.py", r"SESSION_COOKIE_SAMESITE", "HIGH"),
        ("App: session cookie Secure on HTTPS", "app.py", r"SESSION_COOKIE_SECURE", "HIGH"),
        ("App: CSRF protection registered", "app.py", r"CSRFProtect\(app\)", "HIGH"),
        ("App: security middleware registered", "app.py", r"init_security_middleware\(app\)", "HIGH"),
        ("App: traffic logging registered", "app.py", r"init_traffic_logging\(app\)", "LOW"),
        ("App: CSP middleware applied", "app.py", r"apply_csp_middleware\(app\)", "MEDIUM"),
        ("App: 404 probe tracking active", "app.py", r"Error404Tracker", "MEDIUM"),
        ("App: session cleanup teardown", "app.py", r"teardown_appcontext", "MEDIUM"),
        ("App: sensitive-data log filter", "utils/logging.py", r"SensitiveDataFilter", "HIGH"),
        ("App: single gunicorn worker documented", "start.sh", r"--workers 1", "LOW"),
        ("App: ZMQ bound to loopback", "start.sh", r"ZMQ_HOST = '127\.0\.0\.1'", "HIGH"),
    ]
    for name, path, pat, sev in checks:
        p = REPO / path
        ok = p.exists() and re.search(pat, p.read_text(errors="replace"))
        add(name, sev, "PASS" if ok else "FAIL",
            f"{'present' if ok else 'NOT FOUND'} in {path}",
            "" if ok else f"This control appears to have been removed from {path} -- investigate")


# ==========================================================================
# 7. Dependencies and static analysis
# ==========================================================================
def check_pip_audit():
    rc, out, err = run(["uv", "run", "--group", "dev", "pip-audit", "--format", "json"])
    if rc < 0:
        add("Dependencies: no known CVEs", "HIGH", "ERROR", f"could not run ({err})")
        return
    try:
        data = json.loads(out)
    except json.JSONDecodeError:
        add("Dependencies: no known CVEs", "HIGH", "ERROR", "unparseable output")
        return
    deps = data.get("dependencies", data) if isinstance(data, dict) else data
    vulns = [(d["name"], d["version"], v) for d in deps for v in d.get("vulns", [])]
    if vulns:
        add("Dependencies: no known CVEs", "HIGH", "FAIL",
            f"{len(vulns)} vulnerable dependencies of {len(deps)} packages",
            "Bump pinned versions then uv sync to regenerate uv.lock",
            "; ".join(f"{n} {v} {x['id']}" for n, v, x in vulns[:6]))
    else:
        add("Dependencies: no known CVEs", "HIGH", "PASS",
            f"{len(deps)} packages scanned, 0 vulnerable")


def check_bandit():
    # Scratch output goes to the OS temp dir, never tmp/ -- tmp/ holds reports only.
    with tempfile.TemporaryDirectory(prefix="openalgo-audit-") as td:
        out_file = Path(td) / "bandit.json"
        run(["uv", "run", "--group", "dev", "bandit", "-r", ".", "-x",
             "./.venv,./frontend,./node_modules,./tmp,./db,./log", "-f", "json",
             "-o", str(out_file), "-q"])
        if not out_file.exists():
            add("Static analysis: bandit", "MEDIUM", "ERROR", "could not run",
                "uv sync --group dev")
            return []
        results = json.loads(out_file.read_text()).get("results", [])
    counts = Counter((r["issue_severity"], r["issue_confidence"]) for r in results)
    actionable = [r for r in results
                  if r["issue_severity"] == "HIGH"
                  or (r["issue_severity"] == "MEDIUM" and r["issue_confidence"] == "HIGH")]
    deployed = [r for r in actionable
                if not re.match(r"^\./(examples|scripts|test|upgrade)/", r["filename"])]
    add("Static analysis: bandit triage", "MEDIUM", "REVIEW" if deployed else "PASS",
        f"{len(results)} raw -> {len(actionable)} actionable -> {len(deployed)} in deployed code",
        "Raw count is not a metric -- most LOW findings are asserts and try/except/pass. "
        "Review the 'Bandit (triaged)' sheet only.",
        "; ".join(f"{k[0]}/{k[1]}={v}" for k, v in sorted(counts.items())))
    return deployed


# ==========================================================================
# Report
# ==========================================================================
SEV_ORDER = {"CRITICAL": 0, "HIGH": 1, "MEDIUM": 2, "LOW": 3}
STATUS_ORDER = {"FAIL": 0, "ERROR": 1, "REVIEW": 2, "WARN": 3, "PASS": 4, "SKIP": 5}

MANUAL_CHECKS = [
    ["Broker credentials", "Confirm the broker portal still whitelists ONLY this server's static IP",
     "SEBI mandate (Apr 2026): stolen keys are unusable off-IP, but anything routed THROUGH this server is not"],
    ["API keys", "Review /apikey; revoke keys for integrations no longer in use",
     "TradingView/Chartink send keys in the body or query string and cannot set headers -- a stale key is a standing grant"],
    ["Sessions", "Check active sessions; confirm every logged-in device is yours (cap 5)",
     "Multi-device login is legitimate, so an extra session is not self-evidently an intrusion"],
    ["Access logs", "Review /traffic and /security for unfamiliar IPs and 404 bursts",
     "Error404Tracker auto-bans probing IPs; a spike indicates active scanning"],
    ["Error log", "Skim log/errors.jsonl for auth failures and decrypt errors",
     "Repeated Fernet decrypt failures can indicate a tampered .env or swapped database"],
    ["Network exposure", "Confirm 5000/8765/5555 are not reachable from the internet",
     "ZMQ 5555 must stay loopback -- it carries the raw tick feed with no authentication"],
    ["Upgrades", "Compare deployed version against the latest release",
     "Security fixes ship in point releases; utils/version.py is the source of truth"],
    ["Backups", "Verify db/ backups exist, restore cleanly, and are encrypted at rest",
     "db/openalgo.db holds Argon2 password hashes and Fernet-encrypted broker tokens"],
    ["Pepper discipline", "Confirm nobody hand-edited API_KEY_PEPPER on a populated database",
     "Rotating the pepper invalidates every password hash and encrypted token; use upgrade/rotate_pepper.py"],
    ["Dependency review", "Skim uv.lock diffs since the last audit for unexpected packages",
     "pip-audit catches known CVEs only -- not a malicious or typosquatted new dependency"],
    ["Sandbox isolation", "Confirm analyzer/sandbox mode cannot reach a live broker",
     "sandbox.db is a separate database; a leak here means simulated orders hit the market"],
    ["Webhook tokens", "Rotate Flow/strategy webhook tokens; confirm none are in shared docs",
     "Webhook URLs are bearer credentials -- anyone with the URL can trigger the strategy"],
]


def write_xlsx(bandit_rows, code_rows, route_rows, fe_rows):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    hdr_font, hdr_fill = Font(bold=True, color="FFFFFF"), PatternFill("solid", fgColor="1F3864")
    fills = {"FAIL": PatternFill("solid", fgColor="F8CBAD"),
             "ERROR": PatternFill("solid", fgColor="F8CBAD"),
             "REVIEW": PatternFill("solid", fgColor="FFE699"),
             "WARN": PatternFill("solid", fgColor="FFF2CC"),
             "PASS": PatternFill("solid", fgColor="E2EFDA"),
             "SKIP": PatternFill("solid", fgColor="EDEDED")}

    def sheet(ws, headers, rows, widths, status_col=None):
        ws.append(headers)
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=c)
            cell.font, cell.fill = hdr_font, hdr_fill
        for r in rows:
            ws.append(r)
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        for row in ws.iter_rows(min_row=2):
            if status_col:
                f = fills.get(str(row[status_col - 1].value))
                if f:
                    for cell in row:
                        cell.fill = f
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        ws.freeze_panes = "A2"
        if ws.max_row > 1:
            ws.auto_filter.ref = ws.dimensions

    ordered = sorted(findings, key=lambda f: (STATUS_ORDER.get(f["Status"], 9),
                                              SEV_ORDER.get(f["Severity"], 9)))
    by_status = Counter(f["Status"] for f in findings)
    open_hi = [f for f in findings if f["Status"] in ("FAIL", "ERROR")
               and f["Severity"] in ("CRITICAL", "HIGH")]

    ws = wb.active
    ws.title = "Summary"
    for row in ([["OpenAlgo Security Audit", ""],
                 ["Generated (UTC)", datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S")],
                 ["Repository", str(REPO)],
                 ["Git commit", run(["git", "rev-parse", "--short", "HEAD"])[1].strip()],
                 ["Git branch", run(["git", "rev-parse", "--abbrev-ref", "HEAD"])[1].strip()],
                 ["", ""],
                 ["Total checks", len(findings)],
                 ["CRITICAL/HIGH open", len(open_hi)],
                 ["", ""], ["Status", "Count"]]
                + [[k, by_status[k]] for k in STATUS_ORDER if by_status.get(k)]):
        ws.append(row)
    ws["A1"].font = Font(bold=True, size=14)
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 64

    sheet(wb.create_sheet("Findings"),
          ["Check", "Severity", "Status", "Detail", "Remediation", "Evidence"],
          [[f["Check"], f["Severity"], f["Status"], f["Detail"], f["Remediation"], f["Evidence"]]
           for f in ordered], [42, 11, 10, 62, 64, 58], status_col=3)

    sheet(wb.create_sheet("Code patterns"),
          ["Severity", "Pattern", "File", "Line", "Source"],
          sorted(code_rows, key=lambda r: SEV_ORDER.get(r[0], 9)), [11, 34, 46, 8, 96])

    sheet(wb.create_sheet("Frontend patterns"),
          ["Severity", "Pattern", "File", "Line", "Source"],
          sorted(fe_rows, key=lambda r: SEV_ORDER.get(r[0], 9)), [11, 36, 50, 8, 90])

    sheet(wb.create_sheet("Routes (unprotected)"),
          ["File", "Line", "Path", "Methods", "Handler"],
          [[r["file"], r["line"], r["path"], r["methods"], r["handler"]] for r in route_rows],
          [30, 8, 46, 20, 52])

    sheet(wb.create_sheet("Bandit (triaged)"),
          ["Severity", "Confidence", "Test", "File", "Line", "Issue"],
          [[r["issue_severity"], r["issue_confidence"], r["test_id"], r["filename"],
            r["line_number"], r["issue_text"]] for r in bandit_rows], [11, 12, 9, 50, 8, 84])

    sheet(wb.create_sheet("Manual review"),
          ["Area", "What to verify", "Why it matters"], MANUAL_CHECKS, [24, 74, 82])

    wb.save(REPORT)


def main():
    print(f"OpenAlgo security audit -- {TODAY}")
    OUTDIR.mkdir(parents=True, exist_ok=True)
    env = parse_env()

    print("  [1/9] secrets and key hygiene")
    check_keys(env); check_secrets_not_committed()
    print("  [2/9] runtime posture")
    check_runtime_config(env)
    print("  [3/9] route protection and CSRF exemptions")
    route_rows = check_route_protection(); check_dev_routes(); check_csrf_exemptions()
    print("  [4/9] auth, headers, strategy sandbox, websocket, container")
    check_auth_hardening(); check_security_headers(); check_strategy_sandbox()
    check_websocket_auth(); check_container_hardening(); check_logging_hygiene()
    check_dependency_integrity()
    print("  [5/9] database and cache security")
    check_database(env); check_cache()
    print("  [6/9] application invariants")
    check_app_invariants()
    print("  [7/9] code vulnerability patterns (backend + frontend)")
    code_rows = check_code_patterns(); fe_rows = check_frontend()
    print("  [8/9] dependency CVEs (slow)")
    check_pip_audit()
    print("  [9/9] static analysis, secret scan, git history (slow)")
    bandit_rows = check_bandit(); check_detect_secrets(); check_git_history_secrets()

    write_xlsx(bandit_rows, code_rows, route_rows, fe_rows)

    by_status = Counter(f["Status"] for f in findings)
    open_hi = [f for f in findings if f["Status"] in ("FAIL", "ERROR")
               and f["Severity"] in ("CRITICAL", "HIGH")]
    print(f"\nReport: {REPORT}")
    print("  " + "  ".join(f"{k}={v}" for k, v in by_status.items()))
    review = [f for f in findings if f["Status"] == "REVIEW"]
    if open_hi:
        print(f"\n{len(open_hi)} CRITICAL/HIGH failing:")
        for f in open_hi:
            print(f"  [{f['Severity']}] {f['Check']}: {f['Detail']}")
    if review:
        print(f"\n{len(review)} item(s) need human triage:")
        for f in review:
            print(f"  [{f['Severity']}] {f['Check']}: {f['Detail']}")
    if not open_hi:
        print("\nNo open CRITICAL/HIGH findings.")
    return 1 if open_hi else 0


if __name__ == "__main__":
    sys.exit(main())
